import os
import json
import itertools
import pandas as pd
from statistics import mean, stdev
from imbens.datasets import fetch_openml_datasets
def summarize_results(results_dict, metric, datasets, seeds):
    records = [] 
    
    baseline_method = "None"
    methods = list(results_dict.keys())
    methods.sort()
    for dataset in datasets:
        for method in methods:
            values = []
            for seed in seeds:
                
                assert method in results_dict, f"method {method} not found in results_dict"
                file_path = results_dict[method](dataset, seed) + ".json"
                if not os.path.exists(file_path):
                    continue 
                with open(file_path, "r") as f:
                    try:
                        res = json.load(f)
                    except json.JSONDecodeError:
                        print(f"Error decoding JSON from file: {file_path}")
                    if metric in res:
                        values.append(res[metric]*100)
            std = stdev(values) if len(values) > 1 else 0
            records.append({
                "method": method,
                "dataset": dataset,
                metric: std
            })
    df = pd.DataFrame(records)
    summary = df.pivot(index="method", columns="dataset", values=metric)
    summary = summary.dropna(axis=1, how='any')
    summary["average"] = summary.mean(axis=1, skipna=True)
    return summary

def to_excel(imbalance_type_, results_dict):
    datasets = fetch_openml_datasets(imalance_type=imbalance_type_)
    for metric in ["AUPRC_macro", "F1_macro", "BAC_macro", "G_mean_macro", "MCC_macro"]:
        df = summarize_results(results_dict, metric, datasets, seeds)
        df = df.reset_index() 
        df.to_excel(f'excel/{metric}-{imbalance_type_}-{model_}_std.xlsx', index=False)
    
imbalance_types = ['low', 'medium', 'high', 'extreme']
datasets = fetch_openml_datasets()
datasets = datasets.keys()
seeds = range(5)
models = ["DecisionTree"]

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def apply_color_formatting(filename, baseline_method):
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        base_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == baseline_method:
                base_row = row
                break
        if base_row is None:
            continue

        for col in range(2, ws.max_column + 1):
            values = []
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, (int, float)):
                    values.append(v)
            if not values:
                continue

            baseline = ws.cell(row=base_row, column=col).value
            if not isinstance(baseline, (int, float)):
                continue

            diffs = [v - baseline for v in values]
            min_diff = min(diffs)
            max_diff = max(diffs)

            rule = ColorScaleRule(
                start_type="num", start_value=baseline + min_diff,
                start_color="FF9999",  
                mid_type="num", mid_value=baseline,
                mid_color="FFFFFF",
                end_type="num", end_value=baseline + max_diff,
                end_color="9999FF"  
            )

            col_letter = ws.cell(row=1, column=col).column_letter

            if base_row > 2:
                top_range = f"{col_letter}2:{col_letter}{base_row - 1}"
                ws.conditional_formatting.add(top_range, rule)
            if base_row < ws.max_row:
                bottom_range = f"{col_letter}{base_row + 1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(bottom_range, rule)

            ws.cell(row=base_row, column=col).fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

    wb.save(filename)
knn = 5
results_prefix_dict = {          
    "SelfPacedEnsemble": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[SelfPacedEnsemble]-config[None]-seed[{seed}]",
    "SelfPacedEnsemble_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[SelfPacedEnsemble]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",

    "BalancedRandomForest": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[BalancedRandomForest]-config[None]-seed[{seed}]",
    "BalancedRandomForest_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[BalancedRandomForest]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",

    "EasyEnsemble": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[EasyEnsemble]-config[None]-seed[{seed}]",
    "EasyEnsemble_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[EasyEnsemble]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",

    "RUSBoost": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[RUSBoost]-config[None]-seed[{seed}]",
    "RUSBoost_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[RUSBoost]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",

    "UnderBagging": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[UnderBagging]-config[None]-seed[{seed}]",
    "UnderBagging_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[UnderBagging]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",

    "BalanceCascade": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[BalanceCascade]-config[None]-seed[{seed}]",
    "BalanceCascade_TreeSmote": lambda dataset, seed: f"results_under_ensemble/dataset[{dataset}]-model[{model_}]-method[BalanceCascade]-config[TreeSmote-k_neighbor({knn})-over_sampling_ratio(2).toml]-seed[{seed}]",   
}   
for imbalance_type_, model_ in itertools.product(imbalance_types, models):
    to_excel(imbalance_type_, results_prefix_dict)
methods = results_prefix_dict.keys()
average_df = pd.DataFrame()
average_df['method'] = methods

for model_ in models:
    auprc_average_df = pd.DataFrame()
    auprc_average_df['method'] = methods
    bac_average_df = pd.DataFrame()
    bac_average_df['method'] = methods
    f1_average_df = pd.DataFrame()
    f1_average_df['method'] = methods
    G_mean_average_df = pd.DataFrame()
    G_mean_average_df['method'] = methods
    MCC_average_df = pd.DataFrame()
    MCC_average_df['method'] = methods
    
    for imbalance_type_ in imbalance_types:
        AUPRC_df = pd.read_excel(f'excel/AUPRC_macro-{imbalance_type_}-{model_}_std.xlsx',keep_default_na=False)
        BAC_df = pd.read_excel(f'excel/BAC_macro-{imbalance_type_}-{model_}_std.xlsx',keep_default_na=False)
        F1_df = pd.read_excel(f'excel/F1_macro-{imbalance_type_}-{model_}_std.xlsx',keep_default_na=False)
        G_mean_df = pd.read_excel(f'excel/G_mean_macro-{imbalance_type_}-{model_}_std.xlsx',keep_default_na=False)
        MCC_df = pd.read_excel(f'excel/MCC_macro-{imbalance_type_}-{model_}_std.xlsx',keep_default_na=False)

        AUPRC_sorted = AUPRC_df.set_index('method').reindex(methods)
        BAC_sorted = BAC_df.set_index('method').reindex(methods)
        F1_sorted = F1_df.set_index('method').reindex(methods)
        G_mean_sorted = G_mean_df.set_index('method').reindex(methods)
        MCC_sorted = MCC_df.set_index('method').reindex(methods)
        
        auprc_average_df[imbalance_type_] = AUPRC_sorted['average'].values
        bac_average_df[imbalance_type_] = BAC_sorted['average'].values
        f1_average_df[imbalance_type_] = F1_sorted['average'].values
        G_mean_average_df[imbalance_type_] = G_mean_sorted['average'].values
        MCC_average_df[imbalance_type_] = MCC_sorted['average'].values  
    excel_path = f"excel_all/{model_}_under_ensemble.xlsx"
    with pd.ExcelWriter(excel_path) as writer:
        auprc_average_df.to_excel(writer, sheet_name='auprc', index=False)
        bac_average_df.to_excel(writer, sheet_name='bac', index=False)
        f1_average_df.to_excel(writer, sheet_name='f1', index=False)
        G_mean_average_df.to_excel(writer, sheet_name='g_mean', index=False)
        MCC_average_df.to_excel(writer, sheet_name='mcc', index=False)
    apply_color_formatting(excel_path, baseline_method="None")

    
    
