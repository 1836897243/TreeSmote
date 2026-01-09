import os
import json
import itertools
import pandas as pd
from statistics import mean, stdev
from imbens.datasets import fetch_openml_datasets
def summarize_results(results_dict, metric, datasets, seeds):
    """
    统计所有数据集和method在所有seed下的平均指标结果。

    参数：
    - metric: str, 需要统计的指标名，例如 "AUPRC_macro", "F1_macro", "BAC_macro"
    - model: str, 指定模型，例如 "DecisionTree", "LinearSVC"
    - datasets: list of str, 数据集名称列表
    - methods: list of str, method名称列表
    - seeds: list of int, 所有seed

    返回：
    - pandas.DataFrame, 行是method，列是dataset，值是平均指标
    """

    # 用来存储所有结果
    records = [] #'TreeSMOTE', 'TreeSMOTE2', 'TreeSMOTE3'
    
    baseline_method = "None"
    methods = list(results_dict.keys())
    # sort
    methods.sort()
    for dataset in datasets:
        for method in methods:
            values = []
            for seed in seeds:
                
                assert method in results_dict, f"method {method} not found in results_dict"
                file_path = results_dict[method](dataset, seed) + ".json"
                # sampler = file_path.split('-')[2]
                # if not os.path.exists(f'results/{dataset}-DecisionTree-{sampler}-auto_balance_kn5_TreeSMOTE.toml-{seed}.json'):
                #     import pdb;pdb.set_trace()
                #     continue
                if not os.path.exists(file_path):
                    # print(f'file_path: {file_path} does not exist.')
                    continue  # 如果文件不存在则跳过
                with open(file_path, "r") as f:
                    try:
                        res = json.load(f)
                    except json.JSONDecodeError:
                        print(f"Error decoding JSON from file: {file_path}")
                    if metric in res:
                        values.append(res[metric]*100)
            
            # 计算平均值
            
            avg_value = sum(values) / len(values) if values else None
            records.append({
                "method": method,
                "dataset": dataset,
                metric: avg_value
            })
            
        
    # 转成DataFrame，行是method，列是dataset
    df = pd.DataFrame(records)
    summary = df.pivot(index="method", columns="dataset", values=metric)
    # 移除包含 None 的列（dataset）
    # import pdb; pdb.set_trace()
    summary = summary.dropna(axis=1, how='any')
    # import pdb; pdb.set_trace()
    summary["average"] = summary.mean(axis=1, skipna=True)
    return summary

def to_excel(imbalance_type_, results_dict):
    datasets = fetch_openml_datasets(imalance_type=imbalance_type_)
    # remove 'MiniBooNE' and 'helena'
    # datasets = [d for d in datasets if d not in ['MiniBooNE', 'helena']]
    for metric in ["AUPRC_macro", "F1_macro", "BAC_macro", "G_mean_macro", "MCC_macro"]:
        df = summarize_results(results_dict, metric, datasets, seeds)
        df = df.reset_index()  # ✅ 把 method 索引恢复为普通列
        df.to_excel(f'excel_ablation/{metric}-{imbalance_type_}-{model_}.xlsx', index=False)
    
imbalance_types = ['low', 'medium', 'high', 'extreme']
datasets = fetch_openml_datasets()
datasets = datasets.keys()
# datasets = [d for d in datasets if d not in ['MiniBooNE', 'helena']]
#methods = ["TreeSMOTE", "TreeSMOTE2", "TreeSMOTE3"]#, "SMOTE", "BorderlineSMOTE", "SVMSMOTE", "None"]#KMeansSMOTE, 'SelfPacedEnsemble', 'BalanceCascade', 'EasyEnsemble', 'RUSBoost', 'UnderBagging'
seeds = range(5)
models = ["DecisionTree"]#, "LinearSVC"]

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def apply_color_formatting(filename, baseline_method):
    """
    对 excel_all/<model_>.xlsx 中每个 sheet 应用条件格式：
    """
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        base_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == baseline_method:
                base_row = row
                break
        if base_row is None:
            continue

        for col in range(2, ws.max_column + 1):
            values = []
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, (int, float)):
                    values.append(v)
            if not values:
                continue

            baseline = ws.cell(row=base_row, column=col).value
            if not isinstance(baseline, (int, float)):
                continue

            diffs = [v - baseline for v in values]
            min_diff = min(diffs)
            max_diff = max(diffs)

            # 三色渐变（红 → 白 → 蓝）
            rule = ColorScaleRule(
                start_type="num", start_value=baseline + min_diff,
                start_color="FF9999",  # 深红
                mid_type="num", mid_value=baseline,
                mid_color="FFFFFF",  # 白
                end_type="num", end_value=baseline + max_diff,
                end_color="9999FF"   # 深蓝
            )

            col_letter = ws.cell(row=1, column=col).column_letter

            if base_row > 2:
                top_range = f"{col_letter}2:{col_letter}{base_row - 1}"
                ws.conditional_formatting.add(top_range, rule)
            if base_row < ws.max_row:
                bottom_range = f"{col_letter}{base_row + 1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(bottom_range, rule)

            ws.cell(row=base_row, column=col).fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

    wb.save(filename)
knn = 5
results_prefix_dict = {                                               
    "SelfPacedEnsemble": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-None-{seed}",
    "SelfPacedEnsemble_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_balance_kn5.toml-{seed}",
    # "SelfPacedEnsemble_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_balance_kn3.toml-{seed}",
    # 'SelfPacedEnsemble_3_balance_kn5': lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-3_balance_kn5.toml-{seed}",
    # 'SelfPacedEnsemble_3_verse_imbalance_kn5': lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-3_verse_imbalance_kn5.toml-{seed}",
    "SelfPacedEnsemble_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "SelfPacedEnsemble_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_balance_kn5_TreeSMOTE2.toml-{seed}",
    #"SelfPacedEnsemble_auto_balance_kn5_TreeSMOTE3": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_balance_kn5_TreeSMOTE3.toml-{seed}",
    #"SelfPacedEnsemble_auto_verse_imbalance": lambda dataset, seed : f"results/{dataset}-{model_}-SelfPacedEnsemble-auto_verse_imbalance_kn5.toml-{seed}",


    "BalancedRandomForest": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-None-{seed}",
    "BalancedRandomForest_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_balance_kn5.toml-{seed}",
    # "BalancedRandomForest_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_balance_kn3.toml-{seed}",
    # "BalancedRandomForest_3_balance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-3_balance_kn5.toml-{seed}",
    # "BalancedRandomForest_3_verse_imbalance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-3_verse_imbalance_kn5.toml-{seed}",
    "BalancedRandomForest_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "BalancedRandomForest_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_balance_kn5_TreeSMOTE2.toml-{seed}",
    #"BalancedRandomForest_auto_balance_kn5_TreeSMOTE3": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_balance_kn5_TreeSMOTE3.toml-{seed}",
            #"BalancedRandomForest_auto_verse_imbalance": lambda dataset, seed : f"results/{dataset}-{model_}-BalancedRandomForest-auto_verse_imbalance_kn5.toml-{seed}",
    "EasyEnsemble": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-None-{seed}",
    "EasyEnsemble_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_balance_kn5.toml-{seed}",
    # "EasyEnsemble_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_balance_kn3.toml-{seed}",
    # "EasyEnsemble_3_balance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-3_balance_kn5.toml-{seed}",
    # "EasyEnsemble_3_verse_imbalance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-3_verse_imbalance_kn5.toml-{seed}",
    "EasyEnsemble_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "EasyEnsemble_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_balance_kn5_TreeSMOTE2.toml-{seed}",
    #"EasyEnsemble_auto_balance_kn5_TreeSMOTE3": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_balance_kn5_TreeSMOTE3.toml-{seed}",
            #"EasyEnsemble_auto_verse_imbalance": lambda dataset, seed : f"results/{dataset}-{model_}-EasyEnsemble-auto_verse_imbalance_kn5.toml-{seed}",
    "RUSBoost": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-None-{seed}",
    "RUSBoost_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_balance_kn5.toml-{seed}",
    # "RUSBoost_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_balance_kn3.toml-{seed}",
    # "RUSBoost_3_balance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-3_balance_kn5.toml-{seed}",
    # "RUSBoost_3_verse_imbalance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-3_verse_imbalance_kn5.toml-{seed}",
    "RUSBoost_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "RUSBoost_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_balance_kn5_TreeSMOTE2.toml-{seed}",
    #"RUSBoost_auto_balance_kn5_TreeSMOTE3": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_balance_kn5_TreeSMOTE3.toml-{seed}",
            #"RUSBoost_auto_verse_imbalance": lambda dataset, seed : f"results/{dataset}-{model_}-RUSBoost-auto_verse_imbalance_kn5.toml-{seed}",
    "UnderBagging": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-None-{seed}",
    "UnderBagging_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_balance_kn5.toml-{seed}",
    # "UnderBagging_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_balance_kn3.toml-{seed}",
    # "UnderBagging_3_balance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-3_balance_kn5.toml-{seed}",
    # "UnderBagging_3_verse_imbalance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-3_verse_imbalance_kn5.toml-{seed}",
    "UnderBagging_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "UnderBagging_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_balance_kn5_TreeSMOTE2.toml-{seed}",
    #"UnderBagging_auto_balance_kn5_TreeSMOTE3": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_balance_kn5_TreeSMOTE3.toml-{seed}",
            #"UnderBagging_auto_verse_imbalance": lambda dataset, seed : f"results/{dataset}-{model_}-UnderBagging-auto_verse_imbalance_kn5.toml-{seed}",
    "BalanceCascade": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-None-{seed}",
    "BalanceCascade_auto_balance": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-auto_balance_kn5.toml-{seed}",
    # "BalanceCascade_auto_balance_kn3": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-auto_balance_kn3.toml-{seed}",
    # "BalanceCascade_3_balance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-3_balance_kn5.toml-{seed}",
    # "BalanceCascade_3_verse_imbalance_kn5": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-3_verse_imbalance_kn5.toml-{seed}",
    "BalanceCascade_auto_balance_kn5_TreeSMOTE": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-auto_balance_kn5_TreeSMOTE.toml-{seed}",
    "BalanceCascade_auto_balance_kn5_TreeSMOTE2": lambda dataset, seed : f"results/{dataset}-{model_}-BalanceCascade-auto_balance_kn5_TreeSMOTE2.toml-{seed}",

}   
for imbalance_type_, model_ in itertools.product(imbalance_types, models):
    to_excel(imbalance_type_, results_prefix_dict)
methods = results_prefix_dict.keys()
average_df = pd.DataFrame()
average_df['method'] = methods

for model_ in models:
    auprc_average_df = pd.DataFrame()
    auprc_average_df['method'] = methods
    bac_average_df = pd.DataFrame()
    bac_average_df['method'] = methods
    f1_average_df = pd.DataFrame()
    f1_average_df['method'] = methods
    G_mean_average_df = pd.DataFrame()
    G_mean_average_df['method'] = methods
    MCC_average_df = pd.DataFrame()
    MCC_average_df['method'] = methods
    
    for imbalance_type_ in imbalance_types:
        # 读取Excel文件
        AUPRC_df = pd.read_excel(f'excel_ablation/AUPRC_macro-{imbalance_type_}-{model_}.xlsx',keep_default_na=False)
        BAC_df = pd.read_excel(f'excel_ablation/BAC_macro-{imbalance_type_}-{model_}.xlsx',keep_default_na=False)
        F1_df = pd.read_excel(f'excel_ablation/F1_macro-{imbalance_type_}-{model_}.xlsx',keep_default_na=False)
        G_mean_df = pd.read_excel(f'excel_ablation/G_mean_macro-{imbalance_type_}-{model_}.xlsx',keep_default_na=False)
        MCC_df = pd.read_excel(f'excel_ablation/MCC_macro-{imbalance_type_}-{model_}.xlsx',keep_default_na=False)

        
        # 确保数据按method列对齐
        AUPRC_sorted = AUPRC_df.set_index('method').reindex(methods)
        BAC_sorted = BAC_df.set_index('method').reindex(methods)
        F1_sorted = F1_df.set_index('method').reindex(methods)
        G_mean_sorted = G_mean_df.set_index('method').reindex(methods)
        MCC_sorted = MCC_df.set_index('method').reindex(methods)
        
        # 将对应行的average值添加到结果DataFrame中
        auprc_average_df[imbalance_type_] = AUPRC_sorted['average'].values
        bac_average_df[imbalance_type_] = BAC_sorted['average'].values
        f1_average_df[imbalance_type_] = F1_sorted['average'].values
        G_mean_average_df[imbalance_type_] = G_mean_sorted['average'].values
        MCC_average_df[imbalance_type_] = MCC_sorted['average'].values
    
    # 保存到同一个Excel文件的不同sheet中
    excel_path = f"excel_all_ablation/{model_}.xlsx"
    with pd.ExcelWriter(excel_path) as writer:
        auprc_average_df.to_excel(writer, sheet_name='auprc', index=False)
        bac_average_df.to_excel(writer, sheet_name='bac', index=False)
        f1_average_df.to_excel(writer, sheet_name='f1', index=False)
        G_mean_average_df.to_excel(writer, sheet_name='g_mean', index=False)
        MCC_average_df.to_excel(writer, sheet_name='mcc', index=False)

    # ✨添加条件格式（红蓝深浅）
    apply_color_formatting(excel_path, baseline_method="None")

    
    
