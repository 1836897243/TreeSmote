import os
import json
import itertools
import pandas as pd
from statistics import mean, stdev
from imbens.datasets import fetch_openml_datasets
import numpy as np
def summarize_results(results_dict, metric, datasets, seeds):
    records = []
    methods = list(results_dict.keys())
    methods.sort()
    for dataset in datasets:
        for method in methods:
            values = []
            for seed in seeds:
                
                assert method in results_dict, f"method {method} not found in results_dict"
                file_path = results_dict[method](dataset, seed) + ".json"
                if not os.path.exists(file_path):
                    print(f'file_path: {file_path} does not exist.')
                    continue  
                with open(file_path, "r") as f:
                    try:
                        res = json.load(f)
                    except json.JSONDecodeError:
                        print(f"Error decoding JSON from file: {file_path}")
                    if metric in res:
                        values.append(res[metric]*100)
                    elif metric in std2key and std2key[metric] in res:
                        values.append(res[std2key[metric]]*100)
                    elif metric == "average" or metric == "average_std":
                            avg_metrics = []
                            for m in true_metrics:
                                if m in res:
                                    avg_metrics.append(res[m]*100)
                                else:
                                    print(f"Warning: metric {m} not found in file {file_path} for method {method} on dataset {dataset} with seed {seed}")
                            if avg_metrics:
                                values.append(mean(avg_metrics))
                            else:
                                print(f"Warning: no valid metrics found to compute average for file {file_path} for method {method} on dataset {dataset} with seed {seed}")
            avg_value = sum(values) / len(values) if values else None
            import numpy as np
            std = np.std(values) if values else None
            records.append({
                "method": method,
                "dataset": dataset,
                metric: std if metric in std2key else avg_value
            })
    df = pd.DataFrame(records)
    summary = df.pivot(index="method", columns="dataset", values=metric)
    summary = summary.dropna(axis=1, how='any')
    summary["average"] = summary.mean(axis=1, skipna=True)
    return summary
true_metrics = ["AUPRC_macro", "BAC_macro", "F1_macro", "G_mean_macro", "MCC_macro"]
true_metrics_with_average = true_metrics + ["average"]
metrics = ["AUPRC_macro", "BAC_macro", "F1_macro", "G_mean_macro", "MCC_macro", "AUPRC_macro_std", "BAC_macro_std", "F1_macro_std", "G_mean_macro_std", "MCC_macro_std"]
metrics_with_average = ["AUPRC_macro", "BAC_macro", "F1_macro", "G_mean_macro", "MCC_macro", "AUPRC_macro_std", "BAC_macro_std", "F1_macro_std", "G_mean_macro_std", "MCC_macro_std", "average", 'average_std']
std2key = {
    "AUPRC_macro_std": "AUPRC_macro",
    "BAC_macro_std": "BAC_macro",
    "F1_macro_std": "F1_macro",
    "G_mean_macro_std": "G_mean_macro",
    "MCC_macro_std": "MCC_macro",
    'average_std': 'average'
}
def to_excel(imbalance_type_, results_dict):
    datasets = fetch_openml_datasets(imalance_type=imbalance_type_)
    datasets = [d for d in datasets if d not in ['helena']]
    for metric in metrics_with_average:
        df = summarize_results(results_dict, metric, datasets, seeds)
        df = df.reset_index() 
        df.to_excel(f'excel/smote_{metric}-{imbalance_type_}.xlsx', index=False)
        rank_df = df.copy()
        rank_df = rank_df.drop(columns=["average"])
        rank_df.iloc[:, 1:] = rank_df.iloc[:, 1:].rank(ascending=False, method='min')
        rank_df["average"] = rank_df.iloc[:, 1:].mean(axis=1)
        rank_df.to_excel(f'excel/smote_ranking_{metric}-{imbalance_type_}.xlsx', index=False)
    
imbalance_types = ['low', 'medium', 'high', 'extreme']
datasets = fetch_openml_datasets()
datasets = datasets.keys()
seeds = range(5)
models = ["DecisionTree"]

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

def apply_color_formatting(filename, baseline_method, biggerBetter=True):
    """
    对 excel_all/<model_>.xlsx 中每个 sheet 应用条件格式：
    """
    wb = load_workbook(filename)
    for ws in wb.worksheets:
        base_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == baseline_method:
                base_row = row
                break
        if base_row is None:
            continue

        for col in range(2, ws.max_column + 1):
            values = []
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, (int, float)):
                    values.append(v)
            if not values:
                continue

            baseline = ws.cell(row=base_row, column=col).value
            if not isinstance(baseline, (int, float)):
                continue

            diffs = [v - baseline for v in values]
            min_diff = min(diffs)
            max_diff = max(diffs)

            
            if biggerBetter:
                rule = ColorScaleRule(
                    start_type="num", start_value=baseline + min_diff,
                    start_color="FF9999", 
                    mid_type="num", mid_value=baseline,
                    mid_color="FFFFFF", 
                    end_type="num", end_value=baseline + max_diff,
                    end_color="9999FF" 
                )
            else:
                rule = ColorScaleRule(
                    start_type="num", start_value=baseline + max_diff,
                    start_color="9999FF",
                    mid_type="num", mid_value=baseline,
                    mid_color="FFFFFF",  
                    end_type="num", end_value=baseline + min_diff,
                    end_color="FF9999" 
                )

            col_letter = ws.cell(row=1, column=col).column_letter

            if base_row > 2:
                top_range = f"{col_letter}2:{col_letter}{base_row - 1}"
                ws.conditional_formatting.add(top_range, rule)
            if base_row < ws.max_row:
                bottom_range = f"{col_letter}{base_row + 1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(bottom_range, rule)

            ws.cell(row=base_row, column=col).fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

    wb.save(filename)
knn = 5
baseline_method = "DecisionTree"
if baseline_method == "DecisionTree":
    results_prefix_dict = {
        "IDENTITY": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-None-k_neighbor(5).toml-{seed}',
        "SMOTE": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-SMOTE-k_neighbor(5).toml-{seed}',
        "BorderlineSMOTE": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-BorderlineSMOTE-k_neighbor(5).toml-{seed}',
        "SVMSMOTE": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-SVMSMOTE-k_neighbor(5).toml-{seed}',
        "SMOTE_TL": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-SMOTE_TL-k_neighbor(5).toml-{seed}',
        "TreeSMOTE": lambda dataset, seed: f'results_smote/{dataset}-DecisionTree-TreeSmote-k_neighbor(5).toml-{seed}',
    }
elif baseline_method == "LinearSVC":
    results_prefix_dict = {
        "IDENTITY": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-None-k_neighbor(5).toml-{seed}',
        "SMOTE": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-SMOTE-k_neighbor(5).toml-{seed}',
        "BorderlineSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-BorderlineSMOTE-k_neighbor(5).toml-{seed}',
        "SVMSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-SVMSMOTE-k_neighbor(5).toml-{seed}',
        "SMOTE_TL": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-SMOTE_TL-k_neighbor(5).toml-{seed}',
        "TreeSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LinearSVC-TreeSmote-k_neighbor(5).toml-{seed}',
    }
elif baseline_method == "MLP":
    results_prefix_dict = {
        "IDENTITY": lambda dataset, seed: f'results_smote/{dataset}-MLP-None-k_neighbor(5).toml-{seed}',
        "SMOTE": lambda dataset, seed: f'results_smote/{dataset}-MLP-SMOTE-k_neighbor(5).toml-{seed}',
        "BorderlineSMOTE": lambda dataset, seed: f'results_smote/{dataset}-MLP-BorderlineSMOTE-k_neighbor(5).toml-{seed}',
        "SVMSMOTE": lambda dataset, seed: f'results_smote/{dataset}-MLP-SVMSMOTE-k_neighbor(5).toml-{seed}',
        "SMOTE_TL": lambda dataset, seed: f'results_smote/{dataset}-MLP-SMOTE_TL-k_neighbor(5).toml-{seed}',
        "TreeSMOTE": lambda dataset, seed: f'results_smote/{dataset}-MLP-TreeSmote-k_neighbor(5).toml-{seed}',
    }
elif baseline_method == "LGBM":
    results_prefix_dict = {
        "IDENTITY": lambda dataset, seed: f'results_smote/{dataset}-LGBM-None-k_neighbor(5).toml-{seed}',
        "SMOTE": lambda dataset, seed: f'results_smote/{dataset}-LGBM-SMOTE-k_neighbor(5).toml-{seed}',
        "BorderlineSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LGBM-BorderlineSMOTE-k_neighbor(5).toml-{seed}',
        "SVMSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LGBM-SVMSMOTE-k_neighbor(5).toml-{seed}',
        "SMOTE_TL": lambda dataset, seed: f'results_smote/{dataset}-LGBM-SMOTE_TL-k_neighbor(5).toml-{seed}',
        "TreeSMOTE": lambda dataset, seed: f'results_smote/{dataset}-LGBM-TreeSmote-k_neighbor(5).toml-{seed}',
    }
else:
    raise ValueError(f"Unsupported baseline method: {baseline_method}")

for imbalance_type_, model_ in itertools.product(imbalance_types, models):
    to_excel(imbalance_type_, results_prefix_dict)
methods = results_prefix_dict.keys()
average_df = pd.DataFrame()
average_df['method'] = methods

excel_path = f"excel_all/smote-{baseline_method}.xlsx"
with pd.ExcelWriter(excel_path) as writer:
    for metric in true_metrics_with_average:
        print(f'processing metric: {metric}')
        average_df = pd.DataFrame()
        average_df['method'] = methods
        sum_df = pd.DataFrame()
        sum_df['method'] = methods
        count = 0
        for imbalance_type_ in imbalance_types:
            df = pd.read_excel(f'excel/smote_{metric}-{imbalance_type_}.xlsx',keep_default_na=False)
            sorted_df = df.set_index('method').reindex(methods)
            average_df[f'{imbalance_type_}'] = sorted_df['average'].values
            sum_df[f'{imbalance_type_}'] = sorted_df['average'].values*len(df.columns[1:-1])
            count += len(df.columns[1:-1])
        try:
            average_df['average'] = sum_df.iloc[:, 1:].sum(axis=1) / count
        except Exception as e:
            print(e)

        std_df = pd.DataFrame()
        std_df['method'] = methods
        sum_df = pd.DataFrame()
        sum_df['method'] = methods
        count = 0
        for imbalance_type_ in imbalance_types:
            df = pd.read_excel(f'excel/smote_{metric}_std-{imbalance_type_}.xlsx',keep_default_na=False)
            sorted_df = df.set_index('method').reindex(methods)
            std_df[f'{imbalance_type_}'] = sorted_df['average'].values
            sum_df[f'{imbalance_type_}'] = sorted_df['average'].values*len(df.columns[1:-1])
            count += len(df.columns[1:-1])
        try:
            std_df['average'] = sum_df.iloc[:, 1:].sum(axis=1) / count
        except Exception as e:
            print(e)
        average_with_std = average_df.copy()
        for col in average_df.columns[1:]:
            average_with_std[col] = average_df[col].map('{:.2f}'.format) + "±" + std_df[col].map('{:.2f}'.format)
        average_with_std.to_excel(writer, sheet_name=metric, index=False)
apply_color_formatting(excel_path, baseline_method=baseline_method)
excel_path = f"excel_all/smote_ranking-{baseline_method}.xlsx"
with pd.ExcelWriter(excel_path) as writer:
    for metric in true_metrics_with_average:
        average_df = pd.DataFrame()
        average_df['method'] = methods
        sum_df = pd.DataFrame()
        sum_df['method'] = methods
        count = 0
        for imbalance_type_ in imbalance_types:
            df = pd.read_excel(f'excel/smote_ranking_{metric}-{imbalance_type_}.xlsx',keep_default_na=False)
            sorted_df = df.set_index('method').reindex(methods)
            average_df[f'{imbalance_type_}'] = sorted_df['average'].values

            sum_df[f'{imbalance_type_}'] = sorted_df['average'].values*len(df.columns[1:-1])
            count += len(df.columns[1:-1])
        try:
            average_df['average'] = sum_df.iloc[:, 1:].sum(axis=1) / count
        except Exception as e:
            print(e)
        average_df.to_excel(writer, sheet_name=metric, index=False)
apply_color_formatting(excel_path, baseline_method=baseline_method, biggerBetter=False)


   